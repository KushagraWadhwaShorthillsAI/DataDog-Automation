#!/usr/bin/env python3
"""
Enhanced script to format daily analysis files with color highlighting and bold formatting
"""

import pandas as pd
import os
import glob
import re
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

def parse_daily_analysis_file(file_path):
    """Parse a daily analysis file and extract metrics in table format"""
    
    with open(file_path, 'r', encoding='utf-8') as f:
        content = f.read()
    
    # Extract service name from directory
    service_name = os.path.basename(os.path.dirname(file_path))
    
    # Extract dates from filename
    filename = os.path.basename(file_path)
    match = re.search(r'daily_analysis_(\d+-\d+)_vs_(\d+-\d+)\.txt', filename)
    if not match:
        return None
    
    date1, date2 = match.groups()
    
    # Extract actual dates from the comparison line in the file
    # Format: Comparison: 2025-10-02 → 2025-10-03
    comparison_match = re.search(r'Comparison:\s+(\d{4}-\d{2}-\d{2})\s+→\s+(\d{4}-\d{2}-\d{2})', content)
    
    # Store full dates for parsing
    full_date1 = None
    full_date2 = None
    
    if comparison_match:
        # Extract just day-month for display in DD-MM format
        full_date1 = comparison_match.group(1)  # Full date format: 2025-10-02
        full_date2 = comparison_match.group(2)  # Full date format: 2025-10-03
        display_date1 = f"{full_date1[-2:]}-{full_date1[-5:-3]}"  # Extract DD-MM: 02-10
        display_date2 = f"{full_date2[-2:]}-{full_date2[-5:-3]}"  # Extract DD-MM: 03-10
        
        # Use these dates instead of the filename dates
        date1, date2 = display_date1, display_date2
    
    # Parse metrics from content
    metrics = {}
    
    # Parse each metric section
    sections = content.split('\n\n')
    
    for section in sections:
        # Add the comparison line to each section to ensure date extraction works
        if comparison_match and 'Comparison:' not in section:
            section = f"Comparison: {full_date1} → {full_date2}\n{section}"
            
        if 'Latency Metric' in section:
            metrics['Latency'] = parse_metric_section(section, date1, date2)
        elif 'Throughput Metric' in section:
            metrics['Throughput'] = parse_metric_section(section, date1, date2)
        elif 'LLM Cost Metric' in section:
            metrics['LLM Cost'] = parse_metric_section(section, date1, date2)
        elif 'Reliability Metric' in section:
            metrics['Reliability'] = parse_metric_section(section, date1, date2)
        elif 'User Activity Metric' in section:
            metrics['User Activity'] = parse_metric_section(section, date1, date2)
    
    return {
        'service': service_name,
        'date1': date1,
        'date2': date2,
        'metrics': metrics
    }

def parse_metric_section(section, date1, date2):
    """Parse a metric section and extract values"""
    
    lines = section.strip().split('\n')
    
    # Extract values from lines with explicit dates
    # The format is now "YYYY-MM-DD Metric Name: value"
    date1_value = None  # Value for the older date
    date2_value = None  # Value for the newer date
    change_text = None
    status = None
    
    # Extract the full date format from the comparison line if available
    full_date1 = None
    full_date2 = None
    
    # First, try to find the comparison line to get the full dates
    for line in section.strip().split('\n'):
        comparison_match = re.search(r'Comparison:\s+(\d{4}-\d{2}-\d{2})\s+→\s+(\d{4}-\d{2}-\d{2})', line)
        if comparison_match:
            full_date1 = comparison_match.group(1)  # Older date
            full_date2 = comparison_match.group(2)  # Newer date
            break
    
    for line in lines:
        # Check for lines with explicit dates
        if full_date1 and full_date1 in line and ":" in line:
            # Extract numeric value (handle different formats)
            if "Cost" in line:
                # For cost: "2025-10-03 Total Cost ($): 0.59"
                match = re.search(r':\s*\$?([\d.]+)', line)
            elif "Rate" in line:
                # For success rate: "2025-10-03 Success Rate: 99.9%"
                match = re.search(r':\s*([\d.]+)', line)
            else:
                # For other metrics: "2025-10-03 Avg Response Time: 1.354ms"
                match = re.search(r':\s*([\d.]+)', line)
            
            if match:
                try:
                    date1_value = round(float(match.group(1)), 2)
                except ValueError:
                    date1_value = match.group(1)
                
        elif full_date2 and full_date2 in line and ":" in line:
            # Extract numeric value (handle different formats)
            if "Cost" in line:
                # For cost: "2025-10-06 Total Cost ($): 0.64"
                match = re.search(r':\s*\$?([\d.]+)', line)
            elif "Rate" in line:
                # For success rate: "2025-10-06 Success Rate: 99.2%"
                match = re.search(r':\s*([\d.]+)', line)
            else:
                # For other metrics: "2025-10-06 Avg Response Time: 1.210ms"
                match = re.search(r':\s*([\d.]+)', line)
            
            if match:
                try:
                    date2_value = round(float(match.group(1)), 2)
                except ValueError:
                    date2_value = match.group(1)
        
        # Fallback to the old format with "Today's" and "Yesterday's"
        elif "Today's" in line and ":" in line:
            # Extract numeric value (handle different formats)
            if "Cost" in line:
                # For cost: "Today's Total Cost ($): 0.59"
                match = re.search(r':\s*\$?([\d.]+)', line)
            elif "Rate" in line:
                # For success rate: "Today's Success Rate: 99.9%"
                match = re.search(r':\s*([\d.]+)', line)
            else:
                # For other metrics: "Today's Avg Response Time: 1.354ms"
                match = re.search(r':\s*([\d.]+)', line)
            
            if match:
                try:
                    date2_value = round(float(match.group(1)), 2)  # Today refers to newer date (date2)
                except ValueError:
                    date2_value = match.group(1)
                
        elif "Yesterday's" in line and ":" in line:
            # Extract numeric value (handle different formats)
            if "Cost" in line:
                # For cost: "Yesterday's Total Cost ($): 0.64"
                match = re.search(r':\s*\$?([\d.]+)', line)
            elif "Rate" in line:
                # For success rate: "Yesterday's Success Rate: 99.2%"
                match = re.search(r':\s*([\d.]+)', line)
            else:
                # For other metrics: "Yesterday's Avg Response Time: 1.210ms"
                match = re.search(r':\s*([\d.]+)', line)
            
            if match:
                try:
                    date1_value = round(float(match.group(1)), 2)  # Yesterday refers to older date (date1)
                except ValueError:
                    date1_value = match.group(1)
                
        elif "Change:" in line or "Change ($):" in line:
            if "Change ($):" in line:
                change_text = line.split('Change ($):')[1].strip()
            else:
                change_text = line.split('Change:')[1].strip()
        elif "Status:" in line:
            status = line.split('Status:')[1].strip()
    
    return {
        'date1_value': date1_value,  # date1 is the older date
        'date2_value': date2_value,  # date2 is the newer date
        'change': change_text,
        'status': status
    }

def get_status_color(status):
    """Get color based on status"""
    if not status:
        return None
    
    status = status.upper()
    
    # Positive statuses - Green
    if status in ['IMPROVING', 'GROWING', 'EFFICIENT']:
        return PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    
    # Negative statuses - Red
    elif status in ['DEGRADING', 'DECLINING', 'EXPENSIVE']:
        return PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    
    # Neutral statuses - Yellow
    elif status == 'STABLE':
        return PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
    
    return None

def get_change_color(change_text, status):
    """Get color based on change direction - should match status color"""
    if not change_text or not status:
        return None
    
    # Use the same color logic as status
    return get_status_color(status)

def create_formatted_excel():
    """Create enhanced Excel with color highlighting and bold formatting"""
    
    base_dir = "/Users/shtlpmac027/Documents/DataDog/individual_analysis"
    daily_files = glob.glob(f"{base_dir}/**/daily_analysis_*.txt", recursive=True)
    
    # Group by date comparison
    date_groups = {}
    
    for file_path in daily_files:
        parsed_data = parse_daily_analysis_file(file_path)
        if parsed_data:
            date_key = f"{parsed_data['date1']}_vs_{parsed_data['date2']}"
            
            if date_key not in date_groups:
                date_groups[date_key] = []
            
            date_groups[date_key].append(parsed_data)
    
    # Create Excel file with current month name
    from datetime import datetime
    current_month = datetime.now().strftime('%B')
    output_file = f"/Users/shtlpmac027/Documents/DataDog/{current_month}_daily.xlsx"
    
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    # Create index sheet first (will be the first sheet)
    index_sheet = wb.create_sheet(title="Link to other tabs")
    
    # Sort date_groups by date in ascending order
    sorted_date_keys = sorted(date_groups.keys(), key=lambda x: (
        # Parse the first date (date1) from format "DD-MM" to a sortable value
        int(x.split('_vs_')[0].split('-')[1]),  # Month of date1
        int(x.split('_vs_')[0].split('-')[0]),  # Day of date1
        # Then by the second date (date2)
        int(x.split('_vs_')[1].split('-')[1]),  # Month of date2
        int(x.split('_vs_')[1].split('-')[0])   # Day of date2
    ))
    
    # Define styles
    bold_font = Font(bold=True, size=12)
    header_font = Font(bold=True, size=11)
    service_font = Font(bold=True, size=12, color='2F4F4F')
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_alignment = Alignment(horizontal='center', vertical='center')
    
    # Process date groups in sorted order
    for date_key in sorted_date_keys:
        services_data = date_groups[date_key]
        print(f"Processing date comparison: {date_key}")
        
        # Create new worksheet with a cleaner sheet name format
        sheet_name = f"Daily_Analysis_{date_key.replace('-', '_').replace('_vs_', '_vs_')}"
        ws = wb.create_sheet(title=sheet_name)
        
        # Headers - Use the actual dates from the first service in this group
        # Get the dates from the first service in this comparison group
        first_service_date1 = services_data[0]['date1'] if services_data else date_key.split('_vs_')[0]
        first_service_date2 = services_data[0]['date2'] if services_data else date_key.split('_vs_')[1]
        headers = ['Service', f'{first_service_date1}', f'{first_service_date2}', 'Change', 'Status']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            cell.font = Font(bold=True, color='FFFFFF')
            cell.alignment = center_alignment
            cell.border = border
        
        current_row = 2
        
        for service_data in services_data:
            service_name = service_data['service']
            # Get the dates for this specific service
            service_date1 = service_data['date1']
            service_date2 = service_data['date2']
            metrics = service_data['metrics']
            
            # Add service header row
            service_cell = ws.cell(row=current_row, column=1, value=service_name)
            service_cell.font = service_font
            service_cell.fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
            service_cell.alignment = center_alignment
            service_cell.border = border
            
            # Merge cells for service name
            ws.merge_cells(f'A{current_row}:E{current_row}')
            
            current_row += 1
            
            # Add metrics for this service
            for metric_name, metric_data in metrics.items():
                # Service column
                metric_cell = ws.cell(row=current_row, column=1, value=f'{metric_name} Metric')
                metric_cell.font = Font(bold=True)
                metric_cell.border = border
                
                # Date 1 value
                date1_value = metric_data['date1_value']
                if isinstance(date1_value, (int, float)):
                    date1_cell = ws.cell(row=current_row, column=2, value=date1_value)
                    date1_cell.number_format = '0.00'
                else:
                    date1_cell = ws.cell(row=current_row, column=2, value=date1_value or '')
                date1_cell.border = border
                date1_cell.alignment = Alignment(horizontal='right', vertical='center')
                
                # Date 2 value
                date2_value = metric_data['date2_value']
                if isinstance(date2_value, (int, float)):
                    date2_cell = ws.cell(row=current_row, column=3, value=date2_value)
                    date2_cell.number_format = '0.00'
                else:
                    date2_cell = ws.cell(row=current_row, column=3, value=date2_value or '')
                date2_cell.border = border
                date2_cell.alignment = Alignment(horizontal='right', vertical='center')
                
                # Change column with color highlighting
                change_cell = ws.cell(row=current_row, column=4, value=metric_data['change'] or '')
                change_cell.border = border
                change_cell.alignment = Alignment(horizontal='right', vertical='center')
                change_fill = get_change_color(metric_data['change'], metric_data['status'])
                if change_fill:
                    change_cell.fill = change_fill
                
                # Status column with color highlighting
                status_cell = ws.cell(row=current_row, column=5, value=metric_data['status'] or '')
                status_cell.border = border
                status_cell.alignment = Alignment(horizontal='right', vertical='center')
                status_fill = get_status_color(metric_data['status'])
                if status_fill:
                    status_cell.fill = status_fill
                
                current_row += 1
            
            # Add empty row between services
            current_row += 1
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        print(f"✅ Created sheet: Daily_Analysis_{date_key.replace('-', '_')} with {current_row-1} rows")
    
    # Create index sheet content after all other sheets are created
    create_index_sheet(wb, index_sheet)
    
    # Save the workbook
    wb.save(output_file)
    print(f"\n🎉 Enhanced daily analysis created: {current_month}_daily.xlsx")

def create_index_sheet(wb, index_sheet):
    """Create an index sheet with hyperlinks to date comparison sheets and metric definitions"""
    # Title styling
    title_cell = index_sheet.cell(row=1, column=1, value='Daily Analysis Report')
    title_cell.font = Font(bold=True, size=16, color='2F4F4F')
    title_cell.alignment = Alignment(horizontal='center')
    
    subtitle_cell = index_sheet.cell(row=2, column=1, value='Click on any link below to jump to that date comparison:')
    subtitle_cell.font = Font(size=12, italic=True, color='696969')
    
    # Add hyperlinks to each date comparison sheet in chronological order
    row = 4
    # Get sheet names excluding the index sheet, and sort them
    sheet_names = [sheet for sheet in wb.sheetnames if sheet != "Link to other tabs"]
    
    # Extract date parts from the sheet name format "Daily_Analysis_DD_MM_vs_DD_MM"
    def extract_date_parts(sheet_name):
        if not sheet_name.startswith("Daily_Analysis_"):
            return (0, 0, 0, 0)  # Default for non-matching sheets
        
        parts = sheet_name.replace("Daily_Analysis_", "").split("_vs_")
        if len(parts) != 2:
            return (0, 0, 0, 0)
            
        date1_parts = parts[0].split("_")
        date2_parts = parts[1].split("_")
        
        if len(date1_parts) != 2 or len(date2_parts) != 2:
            return (0, 0, 0, 0)
            
        try:
            # Format is DD_MM_vs_DD_MM
            day1 = int(date1_parts[0])
            month1 = int(date1_parts[1])
            day2 = int(date2_parts[0])
            month2 = int(date2_parts[1])
            return (month1, day1, month2, day2)
        except (ValueError, IndexError):
            return (0, 0, 0, 0)
    
    # Sort sheets by date in chronological order
    sorted_sheet_names = sorted(sheet_names, key=extract_date_parts)
    
    for sheet in sorted_sheet_names:
        cell = index_sheet.cell(row=row, column=1)
        cell.value = f"=HYPERLINK(\"#'{sheet}'!A1\",\"{sheet}\")"
        cell.font = Font(size=11, color='0066CC', underline='single')
        cell.alignment = Alignment(horizontal='left')
        row += 1
    
    # Add metric definitions section
    row += 2
    index_sheet.cell(row=row, column=1, value='Metric Definitions').font = Font(bold=True, size=14)
    row += 2
    
    # Get sample dates from the first sheet name for the examples
    sample_newer_date = "06-10"  # Default
    sample_older_date = "03-10"  # Default
    
    # Try to extract dates from the first sheet name if available
    if sorted_sheet_names:
        first_sheet = sorted_sheet_names[0]
        if first_sheet.startswith("Daily_Analysis_"):
            parts = first_sheet.replace("Daily_Analysis_", "").split("_vs_")
            if len(parts) == 2:
                try:
                    older_parts = parts[0].split("_")
                    newer_parts = parts[1].split("_")
                    if len(older_parts) == 2 and len(newer_parts) == 2:
                        sample_older_date = f"{older_parts[0]}-{older_parts[1]}"
                        sample_newer_date = f"{newer_parts[0]}-{newer_parts[1]}"
                except:
                    pass  # Use defaults if any error occurs
    
    # 1. Latency Metric
    index_sheet.cell(row=row, column=1, value='1. Latency Metric').font = Font(bold=True)
    row += 1
    index_sheet.cell(row=row, column=1, value="Definition: Shows the comparison of average response time between two dates in seconds.")
    row += 1
    index_sheet.cell(row=row, column=1, value="Reveals how system performance has changed over time. A decrease in response time indicates improved performance.")
    row += 1
    index_sheet.cell(row=row, column=1, value=f"Example: {sample_newer_date} Avg Response Time: 39.57s, {sample_older_date}: 38.98s, Change: +0.59s (↑1.5% increase)")
    row += 1
    index_sheet.cell(row=row, column=1, value="Status: IMPROVING (response time decreased from older to newer date), DEGRADING (increased), STABLE (minimal change)")
    row += 2
    
    # 2. Throughput Metric
    index_sheet.cell(row=row, column=1, value='2. Throughput Metric').font = Font(bold=True)
    row += 1
    index_sheet.cell(row=row, column=1, value="Definition: Shows the comparison of total request volume between two dates.")
    row += 1
    index_sheet.cell(row=row, column=1, value="Highlights changes in system usage and demand between the compared dates.")
    row += 1
    index_sheet.cell(row=row, column=1, value=f"Example: {sample_newer_date} Total Requests: 1,247, {sample_older_date}: 1,156, Change: +91 requests (↑7.9% increase)")
    row += 1
    index_sheet.cell(row=row, column=1, value="Status: GROWING (requests increased from older to newer date), DECLINING (decreased), STABLE (similar volume)")
    row += 2
    
    # 3. LLM Cost Metric
    index_sheet.cell(row=row, column=1, value='3. LLM Cost Metric').font = Font(bold=True)
    row += 1
    index_sheet.cell(row=row, column=1, value="Definition: Shows the comparison of Large Language Model (LLM) expenditure between two dates.")
    row += 1
    index_sheet.cell(row=row, column=1, value="Tracks how AI processing costs have changed, helping identify cost efficiency trends.")
    row += 1
    index_sheet.cell(row=row, column=1, value=f"Example: {sample_newer_date} Total Cost: $45.67, {sample_older_date}: $42.30, Change: +$3.37 (↑8.0% increase)")
    row += 1
    index_sheet.cell(row=row, column=1, value="Status: EFFICIENT (cost per request decreased from older to newer date), EXPENSIVE (increased), STABLE (similar efficiency)")
    row += 2
    
    # 4. Reliability Metric
    index_sheet.cell(row=row, column=1, value='4. Reliability Metric').font = Font(bold=True)
    row += 1
    index_sheet.cell(row=row, column=1, value="Definition: Shows the comparison of successful request percentages between two dates.")
    row += 1
    index_sheet.cell(row=row, column=1, value="Illustrates how system stability and error rates have evolved between the compared dates.")
    row += 1
    index_sheet.cell(row=row, column=1, value=f"Example: {sample_newer_date} Success Rate: 98.5%, {sample_older_date}: 96.8%, Change: +1.7% (↑1.8% improvement)")
    row += 1
    index_sheet.cell(row=row, column=1, value="Status: IMPROVING (success rate increased from older to newer date), DEGRADING (decreased), STABLE (similar rates)")
    row += 2
    
    # 5. User Activity Metric
    index_sheet.cell(row=row, column=1, value='5. User Activity Metric').font = Font(bold=True)
    row += 1
    index_sheet.cell(row=row, column=1, value="Definition: Shows the comparison of unique user counts between two dates.")
    row += 1
    index_sheet.cell(row=row, column=1, value="Demonstrates how the user base has changed, indicating shifts in adoption and engagement patterns.")
    row += 1
    index_sheet.cell(row=row, column=1, value=f"Example: {sample_newer_date} Unique Users: 892, {sample_older_date}: 847, Change: +45 users (↑5.3% growth)")
    row += 1
    index_sheet.cell(row=row, column=1, value="Status: GROWING (user count increased from older to newer date), DECLINING (decreased), STABLE (similar count)")
    
    # Auto-adjust column width
    index_sheet.column_dimensions['A'].width = 50

if __name__ == "__main__":
    create_formatted_excel()
