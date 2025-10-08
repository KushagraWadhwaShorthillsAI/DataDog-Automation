#!/usr/bin/env python3
"""
Final Polished Combined Report Generator (v3 - Excel Corrected)
Pure regex-based parsing with beautified PDF layouts and a fully corrected,
professionally formatted Excel output.
"""

import os
import sys
import pandas as pd
import matplotlib.pyplot as plt
from matplotlib.backends.backend_pdf import PdfPages
from PIL import Image
from datetime import datetime
import re
import traceback
from typing import Dict, List
import textwrap

# Explicitly import the engine for writing to Excel files for clarity
import openpyxl
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from openpyxl.drawing.image import Image as XLImage
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

# Import LLM service for consistent error categorization
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from llm_service import llm_service

class FinalPolishedCombinedReport:
    """Final combined report generator with pure regex parsing and professional styling"""
    
    def __init__(self):
        # --- Configuration ---
        self.base_dir = "/Users/shtlpmac027/Documents/DataDog"
        self.individual_analysis_dir = f"{self.base_dir}/individual_analysis"
        self.reports_dir = f"{self.base_dir}/combined_reports"
        os.makedirs(self.reports_dir, exist_ok=True)
        
        # --- Professional Styling Configuration ---
        self.A4_SIZE_INCHES = (8.27, 11.69)
        self.FONT_NAME = 'Helvetica'
        plt.rcParams['font.family'] = 'sans-serif'
        plt.rcParams['font.sans-serif'] = self.FONT_NAME
        # PDF page counter for footer
        self._pdf_page_num = 0
    
    def collect_data(self) -> Dict:
        """Collect data from individual analysis folders"""
        print("📊 Collecting individual analysis data...")
        all_data = {}
        for file_dir in sorted(os.listdir(self.individual_analysis_dir)):
            file_path = os.path.join(self.individual_analysis_dir, file_dir)
            if os.path.isdir(file_path):
                metrics_file = os.path.join(file_path, "metrics_analysis.txt")
                if os.path.exists(metrics_file):
                    print(f"  📁 Processing {file_dir}...")
                    metrics = self._parse_metrics_regex_only(metrics_file)
                    # Prefer service name from metrics header if present
                    service_name = metrics.get('__service_display__', file_dir)
                    charts = {}
                    chart_files = [
                        'dau_chart.png', 'dauu_chart.png', 'mode_wise_dau_chart.png', 'response_time_percentiles.png',
                        'response_time_analysis.png', 'daily_response_time_range.png', 'error_categories_chart.png'
                    ]
                    for chart in chart_files:
                        chart_path = os.path.join(file_path, chart)
                        if os.path.exists(chart_path):
                            charts[chart] = chart_path
                    print(f"    📈 Found {len(charts)} charts")
                    all_data[service_name] = {'metrics': metrics, 'charts': charts}
                    print(f"    ✅ Data collected for {service_name}")
        return all_data
    
    def _parse_metrics_regex_only(self, metrics_file: str) -> Dict:
        """Pure regex-based parsing without any LLM usage"""
        with open(metrics_file, 'r', encoding='utf-8') as f:
            content = f.read()
        metrics = {}
        
        try:
            # Capture service display name if emitted by analyzer
            m = re.search(r'^SERVICE NAME:\s*(.+)$', content, re.MULTILINE)
            if m:
                metrics['__service_display__'] = m.group(1).strip()
            
            # Response Time Metrics - with better error handling
            rt_avg = re.search(r'Avg Time Taken \(s\)\s+([0-9.]+)', content)
            if rt_avg:
                try:
                    metrics['response_time'] = {
                        'avg': float(rt_avg.group(1)),
                        'min': float(re.search(r'Min Time Taken \(s\)\s+([0-9.]+)', content).group(1)),
                        'max': float(re.search(r'Max Time Taken \(s\)\s+([0-9.]+)', content).group(1)),
                        'median': float(re.search(r'Median Time \(s\)\s+([0-9.]+)', content).group(1)),
                        'std': float(re.search(r'Std Deviation \(s\)\s+([0-9.]+)', content).group(1)),
                        'count': int(re.search(r'Records Analyzed\s+([0-9,]+)', content).group(1).replace(',', ''))
                    }
                except (AttributeError, ValueError) as e:
                    print(f"⚠️ Error parsing response time metrics: {e}")
            else:
                print(f"⚠️ No response time metrics found in {metrics_file}")
            
            # LLM Cost Metrics - with better error handling
            cost_avg = re.search(r'Avg LLM Cost \(\$\)\s+([0-9.]+)', content)
            if cost_avg:
                try:
                    metrics['llm_cost'] = {
                        'avg': float(cost_avg.group(1)),
                        'min': float(re.search(r'Min LLM Cost \(\$\)\s+([0-9.]+)', content).group(1)),
                        'max': float(re.search(r'Max LLM Cost \(\$\)\s+([0-9.]+)', content).group(1)),
                        'median': float(re.search(r'Median Cost \(\$\)\s+([0-9.]+)', content).group(1)),
                        'total': float(re.search(r'Total LLM Cost \(\$\)\s+([0-9.]+)', content).group(1)),
                        'count': int(re.search(r'Records with Cost\s+([0-9,]+)', content).group(1).replace(',', ''))
                    }
                except (AttributeError, ValueError) as e:
                    print(f"⚠️ Error parsing LLM cost metrics: {e}")
            else:
                print(f"⚠️ No LLM cost metrics found in {metrics_file}")
            
            # Status Metrics - with better error handling
            error_match = re.search(r'error \(Failure\)\s+([\d,]+)\s+([0-9.]+)%', content)
            if error_match:
                try:
                    total_match = re.search(r'Total\s+([\d,]+)\s+100\.00%', content)
                    success_match = re.search(r'info \(Success\)\s+([\d,]+)', content)
                    success_rate_match = re.search(r'info \(Success\)\s+[\d,]+\s+([0-9.]+)%', content)
                    
                    if total_match and success_match and success_rate_match:
                        metrics['status'] = {
                            'total': int(total_match.group(1).replace(',', '')),
                            'success_count': int(success_match.group(1).replace(',', '')),
                            'success_rate': float(success_rate_match.group(1)),
                            'error_count': int(error_match.group(1).replace(',', '')),
                            'error_rate': float(error_match.group(2))
                        }
                except (AttributeError, ValueError) as e:
                    print(f"⚠️ Error parsing status metrics: {e}")
        except Exception as e:
            print(f"❌ Error parsing basic metrics from {metrics_file}: {e}")
            return metrics
        # ERROR MESSAGE TO CATEGORY MAPPING Parsing (Primary source for messages and categories)
        error_message_categories = {}
        try:
            mapping_pattern = r'ERROR MESSAGE TO CATEGORY MAPPING\n=+\n(.*?)\n\nERROR TYPE CATEGORIES'
            mapping_match = re.search(mapping_pattern, content, re.DOTALL)
            if mapping_match:
                for line in mapping_match.group(1).strip().split('\n'):
                    if '|=>|' in line:
                        parts = line.split('|=>|', 1)
                        if len(parts) == 2:
                            category = parts[0].strip()
                            message = parts[1].strip()
                            error_message_categories[message] = category
        except Exception as e:
            print(f"⚠️ Error parsing error message categories: {e}")
        metrics['error_message_categories'] = error_message_categories
        
        # DETAILED ERROR BREAKDOWN Parsing (Get counts for messages)
        error_messages = {}
        full_error_messages = {}  # Store full messages for detailed sheet
        try:
            msg_pattern = r'DETAILED ERROR BREAKDOWN\n=+\nError Message.*?\n-+\n(.*?)\n\nTotal unique error'
            msg_match = re.search(msg_pattern, content, re.DOTALL)
            if msg_match:
                for line in msg_match.group(1).strip().split('\n'):
                    if line.strip():
                        # Split by last occurrence of multiple spaces to separate message from count
                        parts = re.split(r'\s{2,}', line.strip())
                        if len(parts) >= 2 and parts[-1].isdigit():
                            truncated_message = ' '.join(parts[:-1]).strip()
                            count = int(parts[-1])
                            
                            # Find the full message that matches this truncated one
                            full_message = None
                            for full_msg in error_message_categories.keys():
                                if full_msg.startswith(truncated_message) or truncated_message.startswith(full_msg[:50]):
                                    full_message = full_msg
                                    break
                            
                            if full_message:
                                # Aggregate counts for identical messages
                                if full_message in error_messages:
                                    error_messages[full_message] += count
                                    full_error_messages[full_message] += count
                                else:
                                    error_messages[full_message] = count
                                    full_error_messages[full_message] = count
                            else:
                                # If no match found, use truncated message and aggregate
                                if truncated_message in error_messages:
                                    error_messages[truncated_message] += count
                                    full_error_messages[truncated_message] += count
                                else:
                                    error_messages[truncated_message] = count
                                    full_error_messages[truncated_message] = count
        except Exception as e:
            print(f"⚠️ Error parsing detailed error breakdown: {e}")
        metrics['error_messages'] = error_messages
        metrics['full_error_messages'] = full_error_messages  # Store full messages
        
        # ERROR TYPE CATEGORIES Parsing (Category Counts)
        error_categories = {}
        try:
            cat_pattern = r'ERROR TYPE CATEGORIES\n=+\nError Category.*?\n-+\n(.*?)\n\nTotal error categories:'
            cat_match = re.search(cat_pattern, content, re.DOTALL)
            if cat_match:
                for line in cat_match.group(1).strip().split('\n'):
                    if line.strip():
                        # Split by multiple spaces to separate category from count
                        parts = re.split(r'\s{2,}', line.strip())
                        if len(parts) >= 2 and parts[-1].isdigit():
                            category = ' '.join(parts[:-1]).strip()
                            count = int(parts[-1])
                            error_categories[category] = count
        except Exception as e:
            print(f"⚠️ Error parsing error type categories: {e}")
        metrics['error_categories'] = error_categories
        
        # VALIDATION: Cross-check counts and fix discrepancies
        self._validate_and_fix_error_counts(metrics)

        # --- Additional tables: Mode-wise and Process/Mode-wise ---
        def _extract_block(title_regex: str) -> List[str]:
            m = re.search(title_regex, content, re.DOTALL)
            if not m:
                return []
            block = m.group(1).strip()
            return [ln for ln in block.split('\n') if ln.strip()]

        def _split_cols(line: str) -> List[str]:
            return re.split(r'\s{2,}', line.strip())

        def _extract_mode_and_name(cols: List[str]):
            """Handle cases where the first token is 'mode' or 'mode name' is fused as 'mode name'.
            Returns (mode:int, name:str, offset:int) where offset is number of consumed cols for mode+name,
            or None if cannot parse.
            """
            if not cols:
                return None
            c0 = cols[0].strip()
            # Case 1: first col is purely numeric and next col is the name
            if c0.lstrip('-').isdigit() and len(cols) >= 2:
                try:
                    return int(c0), cols[1], 2
                except Exception:
                    return None
            # Case 2: fused token like '7 isDatabaseGeneric'
            m0 = re.match(r'^(\d+)\s+(.+)$', c0)
            if m0:
                try:
                    return int(m0.group(1)), m0.group(2), 1
                except Exception:
                    return None
            return None

        # RESPONSE TIME BY EFFECTIVE MODE (allow optional dashed header line)
        try:
            rt_mode_lines = _extract_block(r'RESPONSE TIME BY EFFECTIVE MODE\n=+\n(?:.*?\n-+\n)?(.*?)\n\n')
            if rt_mode_lines:
                rows = []
                for ln in rt_mode_lines:
                    cols = _split_cols(ln)
                    mn = _extract_mode_and_name(cols)
                    if not mn:
                        continue
                    mode, mode_name, offset = mn
                    # Expect next columns: avg, p50, min, max, std, count
                    if len(cols) >= offset + 6:
                        try:
                            rows.append({
                                'effective_mode': mode,
                                'mode_name': mode_name,
                                'avg': float(cols[offset + 0]),
                                'p50': float(cols[offset + 1]),
                                'min': float(cols[offset + 2]),
                                'max': float(cols[offset + 3]),
                                'std': float(cols[offset + 4]),
                                'count': int(cols[offset + 5])
                            })
                        except (ValueError, IndexError) as e:
                            print(f"⚠️ Error parsing mode response time row: {e}")
                            continue
                if rows:
                    metrics['rt_by_mode'] = rows
        except Exception as e:
            print(f"⚠️ Error parsing response time by mode: {e}")

        # LLM COST BY EFFECTIVE MODE (allow optional dashed header line)
        try:
            cost_mode_lines = _extract_block(r'LLM COST BY EFFECTIVE MODE\n=+\n(?:.*?\n-+\n)?(.*?)\n\n')
            if cost_mode_lines:
                rows = []
                for ln in cost_mode_lines:
                    cols = _split_cols(ln)
                    mn = _extract_mode_and_name(cols)
                    if not mn:
                        continue
                    mode, mode_name, offset = mn
                    # Expect next columns: avg, median, min, max, total, count
                    if len(cols) >= offset + 6:
                        try:
                            rows.append({
                                'effective_mode': mode,
                                'mode_name': mode_name,
                                'avg': float(cols[offset + 0]),
                                'median': float(cols[offset + 1]),
                                'min': float(cols[offset + 2]),
                                'max': float(cols[offset + 3]),
                                'total': float(cols[offset + 4]),
                                'count': int(cols[offset + 5])
                            })
                        except (ValueError, IndexError) as e:
                            print(f"⚠️ Error parsing mode cost row: {e}")
                            continue
                if rows:
                    metrics['cost_by_mode'] = rows
        except Exception as e:
            print(f"⚠️ Error parsing LLM cost by mode: {e}")

        # FAILURE RATE (ERROR COUNTS) BY MODE
        try:
            fail_mode_lines = _extract_block(r'FAILURE RATE \(ERROR COUNTS\) BY MODE\n=+\n(.*?)\n\n')
            if fail_mode_lines:
                rows = []
                for ln in fail_mode_lines:
                    cols = _split_cols(ln)
                    if len(cols) >= 6 and cols[0].strip().lstrip('-').isdigit():
                        try:
                            rows.append({
                                'effective_mode': int(cols[0]),
                                'mode_name': cols[1],
                                'error': int(cols[2]),
                                'info': int(cols[3]),
                                'total': int(cols[4]),
                                'failure_pct': float(cols[5].replace('%','')) / 100.0  # Divide by 100 to convert percentage to decimal
                            })
                        except (ValueError, IndexError) as e:
                            print(f"⚠️ Error parsing mode failure row: {e}")
                            continue
                if rows:
                    metrics['fail_by_mode'] = rows
        except Exception as e:
            print(f"⚠️ Error parsing failure rate by mode: {e}")

        # RESPONSE TIME BY PROCESS
        try:
            rt_proc_lines = _extract_block(r'RESPONSE TIME BY PROCESS\n=+\n.*?\n-+\n(.*?)\n\n')
            if rt_proc_lines:
                rows = []
                for ln in rt_proc_lines:
                    cols = _split_cols(ln)
                    if len(cols) >= 7:
                        try:
                            rows.append({
                                'process_name': cols[0],
                                'avg': float(cols[1]),
                                'p50': float(cols[2]),
                                'min': float(cols[3]),
                                'max': float(cols[4]),
                                'std': float(cols[5]),
                                'count': int(cols[6])
                            })
                        except (ValueError, IndexError) as e:
                            print(f"⚠️ Error parsing process response time row: {e}")
                            continue
                if rows:
                    metrics['rt_by_process'] = rows
        except Exception as e:
            print(f"⚠️ Error parsing response time by process: {e}")

        # LLM COST BY PROCESS
        cost_proc_lines = _extract_block(r'LLM COST BY PROCESS\n=+\n.*?\n-+\n(.*?)\n\n')
        # FAILURE RATE (ERROR COUNTS) BY PROCESS
        # Skip header and dashed line by matching them explicitly before capturing rows
        fail_proc_lines = _extract_block(r'FAILURE RATE \(ERROR COUNTS\) BY PROCESS\n=+\n.*?\n-+\n(.*?)\n\n')
        if fail_proc_lines:
            rows = []
            for ln in fail_proc_lines:
                cols = _split_cols(ln)
                # Ensure this is a data row (not header/overall) by checking numeric columns
                if len(cols) >= 5 and cols[1].replace(',', '').isdigit():
                    rows.append({
                        'process_name': cols[0],
                        'error': int(cols[1]),
                        'info': int(cols[2]),
                        'total': int(cols[3]),
                        'failure_pct': float(cols[4].replace('%','')) / 100.0  # Divide by 100 to convert percentage to decimal
                    })
            metrics['fail_by_process'] = rows
        if cost_proc_lines:
            rows = []
            for ln in cost_proc_lines:
                cols = _split_cols(ln)
                if len(cols) >= 7:
                    rows.append({
                        'process_name': cols[0],
                        'avg': float(cols[1]),
                        'median': float(cols[2]),
                        'min': float(cols[3]),
                        'max': float(cols[4]),
                        'total': float(cols[5]),
                        'count': int(cols[6])
                    })
            metrics['cost_by_process'] = rows

        # RESPONSE TIME BY PROCESS × MODE
        rt_pm_lines = _extract_block(r'RESPONSE TIME BY PROCESS × MODE\n=+\n.*?\n-+\n(.*?)\n\n')
        if rt_pm_lines:
            rows = []
            for ln in rt_pm_lines:
                cols = _split_cols(ln)
                if len(cols) >= 8 and cols[1].strip().lstrip('-').isdigit():
                    rows.append({
                        'process_name': cols[0],
                        'effective_mode': int(cols[1]),
                        'avg': float(cols[2]),
                        'p50': float(cols[3]),
                        'min': float(cols[4]),
                        'max': float(cols[5]),
                        'std': float(cols[6]),
                        'count': int(cols[7])
                    })
            metrics['rt_by_process_mode'] = rows

        # LLM COST BY PROCESS × MODE
        cost_pm_lines = _extract_block(r'LLM COST BY PROCESS × MODE\n=+\n.*?\n-+\n(.*?)\n\n')
        if cost_pm_lines:
            rows = []
            for ln in cost_pm_lines:
                cols = _split_cols(ln)
                if len(cols) >= 8 and cols[1].strip().lstrip('-').isdigit():
                    rows.append({
                        'process_name': cols[0],
                        'effective_mode': int(cols[1]),
                        'avg': float(cols[2]),
                        'median': float(cols[3]),
                        'min': float(cols[4]),
                        'max': float(cols[5]),
                        'total': float(cols[6]),
                        'count': int(cols[7])
                    })
            metrics['cost_by_process_mode'] = rows

        # FAILURE RATE (ERROR COUNTS) BY PROCESS × MODE
        fail_pm_lines = _extract_block(r'FAILURE RATE \(ERROR COUNTS\) BY PROCESS × MODE\n=+\n(.*?)\n\n')
        if fail_pm_lines:
            rows = []
            for ln in fail_pm_lines:
                cols = _split_cols(ln)
                if len(cols) >= 6 and cols[1].strip().lstrip('-').isdigit():
                    rows.append({
                        'process_name': cols[0],
                        'effective_mode': int(cols[1]),
                        'error': int(cols[2]),
                        'info': int(cols[3]),
                        'total': int(cols[4]),
                        'failure_pct': float(cols[5].replace('%','')) / 100.0  # Divide by 100 to convert percentage to decimal
                    })
            metrics['fail_by_process_mode'] = rows
        return metrics
    
    def _validate_and_fix_error_counts(self, metrics: Dict):
        """Validate and fix error count discrepancies between categories and messages."""
        error_categories = metrics.get('error_categories', {})
        error_messages = metrics.get('error_messages', {})
        message_categories = metrics.get('error_message_categories', {})
        
        # Calculate totals
        category_total = sum(error_categories.values())
        message_total = sum(error_messages.values())
        
        print(f"🔍 Validation: Category total={category_total}, Message total={message_total}")
        
        if category_total != message_total:
            print(f"⚠️ Count discrepancy detected: {abs(category_total - message_total)} errors")
            
            # Try to fix by recalculating category counts from messages
            recalculated_categories = {}
            for message, count in error_messages.items():
                category = message_categories.get(message, 'Uncategorized')
                recalculated_categories[category] = recalculated_categories.get(category, 0) + count
            
            # Update metrics with recalculated counts
            metrics['error_categories'] = recalculated_categories
            new_category_total = sum(recalculated_categories.values())
            print(f"✅ Fixed: New category total={new_category_total}, Message total={message_total}")
            
            # Add validation metadata
            metrics['validation'] = {
                'original_category_total': category_total,
                'message_total': message_total,
                'discrepancy': abs(category_total - message_total),
                'fixed': True,
                'recalculated_category_total': new_category_total
            }
        else:
            print("✅ Error counts are consistent")
            metrics['validation'] = {
                'original_category_total': category_total,
                'message_total': message_total,
                'discrepancy': 0,
                'fixed': False
            }
    
    def generate_excel_report(self, all_data: Dict) -> bool:
        """Generate a complete and correctly formatted Excel report."""
        try:
            current_month = datetime.now().strftime('%B')
            excel_path = f"{self.reports_dir}/{current_month}_Complete.xlsx"
            with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
                self._create_response_time_sheet(writer, all_data)
                self._create_success_rate_sheet_restructured(writer, all_data)
                self._create_llm_cost_sheet(writer, all_data)
                # Error Categories table
                self._create_error_categories_sheet(writer, all_data)
                # Add detailed error messages sheet with full text
                self._create_detailed_error_messages_sheet(writer, all_data)
                self._create_charts_sheet(writer, all_data)
                # Per-service consolidated sheets
                self._create_service_sheets(writer, all_data)
                # Index sheet with hyperlinks
                self._create_index_sheet(writer)
            print(f"✅ Excel report: {excel_path}")
            return True
        except Exception as e:
            print(f"❌ Excel generation failed: {e}")
            traceback.print_exc()
            return False
    
    
    def _create_response_time_sheet(self, writer, all_data: Dict):
        """Create a comprehensive response time metrics table for all services"""
        rt_data = []
        for file_name, data in all_data.items():
            rt = data['metrics'].get('response_time')
            if rt:
                rt_data.append([
                    file_name, 
                    rt.get('avg', 0), 
                    rt.get('min', 0), 
                    rt.get('max', 0),
                    rt.get('median', 0), 
                    rt.get('std', 0),
                    rt.get('count', 0)  # Include count for completeness
                ])
        
        if rt_data:
            df = pd.DataFrame(rt_data, columns=[
                'Service', 'Avg Time', 'Min Time', 'Max Time', 
                'Median Time', 'Std Dev', 'Records Analyzed'
            ])
            df.to_excel(writer, sheet_name='Response Times', index=False)
            ws = writer.sheets['Response Times']
            
            # Apply enhanced header styling
            self._apply_header_styling(ws, 1, 1, 7)
            # Apply borders to the entire table
            self._apply_table_borders(ws, 1, len(df) + 1, 1, 7)
            
            # Right-align numeric columns and format
            for row in ws.iter_rows(min_row=2, min_col=2, max_col=7):
                for cell in row:
                    if cell.data_type == 'n':
                        cell.alignment = Alignment(horizontal='right')
                        if cell.column <= 6:  # Time columns
                            cell.number_format = '0.00'  # Remove "s" unit
                        else:  # Count column
                            cell.number_format = '#,##0'
        else:
            # Create empty sheet if no data
            pd.DataFrame(columns=['Service', 'Avg Time', 'Min Time', 'Max Time', 
                                'Median Time', 'Std Dev', 'Records Analyzed']).to_excel(
                writer, sheet_name='Response Times', index=False)
    
    def _create_success_rate_sheet_restructured(self, writer, all_data: Dict):
        """Creates a success rate sheet with true number formatting for percentages."""
        start_row = 0
        for file_name, data in all_data.items():
            st = data['metrics'].get('status', {})
            if st:
                # --- MODIFIED: Write percentages as numbers (e.g., 0.9974) ---
                df = pd.DataFrame({
                    'Status': ['Success', 'Error', 'Total'],
                    'Count': [st.get('success_count', 0), st.get('error_count', 0), st.get('total', 0)],
                    '% of Total': [
                        st.get('success_rate', 0) / 100.0, 
                        st.get('error_rate', 0) / 100.0, 
                        1.0
                    ]
                })
                pd.DataFrame([file_name]).to_excel(writer, sheet_name='Success Rates', startrow=start_row, index=False, header=False)
                df.to_excel(writer, sheet_name='Success Rates', startrow=start_row + 2, index=False)
                
                # Align headers left for this block
                ws = writer.sheets['Success Rates']
                header_row_index = start_row + 3  # because excel is 1-based and header is +2 then +1
                for cell in ws[header_row_index]:
                    cell.alignment = Alignment(horizontal='left'); cell.font = Font(bold=True)

                # --- MODIFIED: Apply percentage number format to the column ---
                worksheet = writer.sheets['Success Rates']
                # The data starts 3 rows down from the start_row (title + blank + header)
                for row_num in range(start_row + 3, start_row + 3 + len(df)):
                    # Right-align numeric columns
                    worksheet[f'B{row_num}'].alignment = Alignment(horizontal='right')
                    worksheet[f'C{row_num}'].alignment = Alignment(horizontal='right')
                    worksheet[f'C{row_num}'].number_format = '0.00%'
                
                start_row += len(df) + 4

    def _create_llm_cost_sheet(self, writer, all_data: Dict):
        cost_data = []
        for file_name, data in all_data.items():
            cost = data['metrics'].get('llm_cost')
            if cost:
                # --- MODIFIED: Removed the 'count' column ---
                cost_data.append([
                    file_name, cost.get('avg', 0), cost.get('min', 0), cost.get('max', 0),
                    cost.get('median', 0), cost.get('total', 0)
                ])
        if cost_data:
            df = pd.DataFrame(cost_data, columns=[
                'File', 'Avg Cost', 'Min Cost', 'Max Cost', 
                'Median Cost', 'Total Cost'
            ])
            df.to_excel(writer, sheet_name='LLM Costs', index=False)
            ws = writer.sheets['LLM Costs']
            
            # Apply enhanced header styling
            self._apply_header_styling(ws, 1, 1, 6)
            # Apply borders to the entire table
            self._apply_table_borders(ws, 1, len(df) + 1, 1, 6)
            
            # Right-align numeric columns and apply number format without currency symbol
            for row in ws.iter_rows(min_row=2, min_col=2, max_col=6):
                for cell in row:
                    if cell.data_type == 'n':
                        cell.alignment = Alignment(horizontal='right')
                        cell.number_format = '#,##0.00'
    
    def _create_error_categories_sheet(self, writer, all_data: Dict):
        """Creates a structured sheet for error categories, grouped by file."""
        start_row = 0
        has_data = False
        for file_name, data in all_data.items():
            error_cats = data['metrics'].get('error_categories', {})
            if error_cats:
                has_data = True
                cat_data = [[category, count] for category, count in error_cats.items()]
                df = pd.DataFrame(cat_data, columns=['Error Category', 'Count'])
                
                pd.DataFrame([file_name]).to_excel(writer, sheet_name='Error Categories', startrow=start_row, index=False, header=False)
                df.to_excel(writer, sheet_name='Error Categories', startrow=start_row + 2, index=False)
                ws = writer.sheets['Error Categories']
                header_row_index = start_row + 3
                for cell in ws[header_row_index]:
                    cell.alignment = Alignment(horizontal='left'); cell.font = Font(bold=True)
                # Right-align numeric counts for this block
                data_start = start_row + 4
                data_end = data_start + len(df) - 1
                for row_num in range(data_start, data_end + 1):
                    ws[f'B{row_num}'].alignment = Alignment(horizontal='right')
                start_row += len(df) + 4
        # If no data was ever written, create an empty sheet to avoid errors
        if not has_data:
            pd.DataFrame().to_excel(writer, sheet_name='Error Categories', index=False)

    def _create_error_messages_sheet(self, writer, all_data: Dict):
        """Creates a structured sheet for error messages, grouped by file."""
        start_row = 0
        has_data = False
        for file_name, data in all_data.items():
            error_msgs = data['metrics'].get('error_messages', {})
            if error_msgs:
                has_data = True
                msg_data = []
                for msg, count in error_msgs.items():
                    display_msg = msg[:300] + "..." if len(msg) > 300 else msg
                    msg_data.append([display_msg, count])
                
                df = pd.DataFrame(msg_data, columns=['Error Message', 'Count'])
                
                pd.DataFrame([file_name]).to_excel(writer, sheet_name='Error Messages', startrow=start_row, index=False, header=False)
                df.to_excel(writer, sheet_name='Error Messages', startrow=start_row + 2, index=False)
                ws = writer.sheets['Error Messages']
                header_row_index = start_row + 3
                for cell in ws[header_row_index]:
                    cell.alignment = Alignment(horizontal='left'); cell.font = Font(bold=True)
                # Right-align numeric counts for this block
                data_start = start_row + 4
                data_end = data_start + len(df) - 1
                for row_num in range(data_start, data_end + 1):
                    ws[f'B{row_num}'].alignment = Alignment(horizontal='right')
                start_row += len(df) + 4
        if not has_data:
            pd.DataFrame().to_excel(writer, sheet_name='Error Messages', index=False)

    # --- New helpers for Category→Message mapping ---
    def _categorize_error_message(self, message: str) -> str:
        """Use the LLM service for consistent error categorization"""
        try:
            return llm_service.categorize_error(message)
        except Exception as e:
            print(f"⚠️ Error categorization failed for message: {e}")
            return 'Other/Uncategorized Errors'


    def _create_detailed_error_messages_sheet(self, writer, all_data: Dict):
        """Create a detailed sheet with full error messages (not truncated)."""
        start_row = 0
        wb = writer.book
        sheet_name = 'Detailed Error Messages'
        ws = wb.create_sheet(sheet_name)
        writer.sheets[sheet_name] = ws
        has_any = False
        
        for file_name, data in all_data.items():
            full_msgs = data['metrics'].get('full_error_messages', {})
            if not full_msgs:
                continue
            has_any = True
            rows = []
            # Use pre-categorized mapping from individual analysis for consistency
            message_categories = data['metrics'].get('error_message_categories', {})
            for msg, count in full_msgs.items():
                # Use pre-categorized mapping if available, otherwise fall back to LLM service
                cat = message_categories.get(msg, self._categorize_error_message(msg))
                rows.append([cat, msg, count])  # Full message, no truncation
            
            # Sort by category then count desc
            df = pd.DataFrame(rows, columns=['Error Category', 'Full Error Message', 'Count'])
            df.sort_values(by=['Error Category', 'Count'], ascending=[True, False], inplace=True)
            
            # Title per service
            pd.DataFrame([file_name]).to_excel(writer, sheet_name=sheet_name, startrow=start_row, index=False, header=False)
            df.to_excel(writer, sheet_name=sheet_name, startrow=start_row + 2, index=False)
            
            # Format block
            header_row_index = start_row + 3
            for cell in ws[header_row_index]:
                cell.alignment = Alignment(horizontal='left'); cell.font = Font(bold=True)
            
            # Right-align counts (third column)
            data_start = start_row + 4
            data_end = data_start + len(df) - 1
            for row_num in range(data_start, data_end + 1):
                ws[f'C{row_num}'].alignment = Alignment(horizontal='right')
            
            # Set column widths for better readability
            ws.column_dimensions['A'].width = 25  # Category
            ws.column_dimensions['B'].width = 100  # Full message
            ws.column_dimensions['C'].width = 10   # Count
            
            start_row += len(df) + 4
        
        if not has_any:
            pd.DataFrame().to_excel(writer, sheet_name=sheet_name, index=False)


    def _create_charts_sheet(self, writer, all_data: Dict):
        """Embed chart images into a Charts sheet in the Excel workbook."""
        wb = writer.book
        ws = wb.create_sheet('Charts')
        current_row = 1
        # Title
        ws.cell(row=current_row, column=1, value='Charts by Service')
        current_row += 2
        # Column A is the anchor for images
        self._charts_anchor_map = {}
        for file_name, data in all_data.items():
            charts = data.get('charts', {})
            if not charts:
                continue
            # Section heading for this service
            ws.cell(row=current_row, column=1, value=f"Service: {file_name}")
            # Remember the first image anchor for hyperlinks
            anchor_row_for_service = current_row + 1
            self._charts_anchor_map[file_name] = f"A{anchor_row_for_service}"
            current_row += 1
            # Keep a consistent order like in PDF
            ordered = [
                'dauu_chart.png',
                'dau_chart.png',
                'mode_wise_dau_chart.png',
                'response_time_percentiles.png',
                'daily_response_time_range.png',
                'response_time_analysis.png',
                'error_categories_chart.png',
            ]
            for chart_file in ordered:
                if chart_file in charts:
                    try:
                        img = XLImage(charts[chart_file])
                        # Scale image to a reasonable width for Excel
                        img.width = 720
                        img.height = 405
                        anchor = f"A{current_row}"
                        ws.add_image(img, anchor)
                        # Advance rows roughly proportional to image height
                        current_row += 28
                    except Exception:
                        # If image fails to load, leave a note
                        ws.cell(row=current_row, column=1, value=f"[Image not found: {charts[chart_file]}]")
                        current_row += 2
            # Gap between different files
            current_row += 2

    def _create_service_sheets(self, writer, all_data: Dict):
        """Create one consolidated sheet per service that includes KPIs, error tables, and charts."""
        wb = writer.book
        self._service_sheet_names: List[str] = []
        for file_name, data in all_data.items():
            # Excel sheet names must be <=31 chars and unique
            base_name = f"{file_name}"
            safe_name = base_name[:31]
            # Ensure uniqueness if truncated duplicates occur
            suffix = 1
            while safe_name in wb.sheetnames:
                candidate = (base_name[:28] + f"-{suffix}")
                safe_name = candidate[:31]
                suffix += 1
            ws = wb.create_sheet(safe_name)
            # Register sheet so pandas writes to it
            writer.sheets[ws.title] = ws
            self._service_sheet_names.append(ws.title)

            current_row = 1
            # Title with enhanced styling
            title_cell = ws.cell(row=current_row, column=1, value=f"Service: {file_name}")
            title_cell.font = Font(bold=True, size=16, color='2F4F4F')
            title_cell.alignment = Alignment(horizontal='center')
            # Add background color to title
            title_cell.fill = PatternFill(start_color='F0F8FF', end_color='F0F8FF', fill_type='solid')
            current_row += 2

            # Separate, neat tables: Success/Error, LLM Cost, Error Categories, Error Messages, then Charts
            # 1) Success/Error table
            st = data['metrics'].get('status', {})
            rt = data['metrics'].get('response_time', {})
            cost = data['metrics'].get('llm_cost', {})
            
            # Add title for Success/Error table
            title_cell = ws.cell(row=current_row, column=1, value="Failure/Success")
            title_cell.font = Font(bold=True, size=12)
            current_row += 1
            
            success_df = pd.DataFrame([
                ['Total', st.get('total', 0)],
                ['Success', st.get('success_count', 0)],
                ['Errors', st.get('error_count', 0)],
                ['Success Rate', (st.get('success_rate', 0) / 100.0) if st else 0.0],
                ['Error Rate', (st.get('error_rate', 0) / 100.0) if st else 0.0],
            ], columns=['Metric', 'Value'])
            success_df.to_excel(writer, sheet_name=ws.title, startrow=current_row-1, index=False)
            header_row = current_row
            # Apply enhanced header styling
            self._apply_header_styling(ws, header_row, 1, 2)
            # Apply borders to the table
            self._apply_table_borders(ws, header_row, header_row + len(success_df), 1, 2)
            
            for r in range(header_row + 1, header_row + 1 + len(success_df)):
                vcell = ws.cell(row=r, column=2)
                if isinstance(vcell.value, (int, float)):
                    vcell.alignment = Alignment(horizontal='right')
            # % format for last two rows
            ws.cell(row=header_row + 4, column=2).number_format = '0.00%'
            ws.cell(row=header_row + 5, column=2).number_format = '0.00%'
            succ_last = header_row + len(success_df)
            current_row = succ_last + 2

            # 2) LLM Cost table
            if cost:
                # Add title for LLM Cost table
                title_cell = ws.cell(row=current_row, column=1, value="LLM Cost ($)")
                title_cell.font = Font(bold=True, size=12)
                current_row += 1
                
                llm_df = pd.DataFrame([
                    ['Avg Cost', cost.get('avg', 0.0)],
                    ['Min Cost', cost.get('min', 0.0)],
                    ['Max Cost', cost.get('max', 0.0)],
                    ['Median Cost', cost.get('median', 0.0)],
                    ['Total Cost', cost.get('total', 0.0)],
                ], columns=['Metric', 'Value'])
                llm_df.to_excel(writer, sheet_name=ws.title, startrow=current_row-1, index=False)
                header_row = current_row
                # Apply enhanced header styling
                self._apply_header_styling(ws, header_row, 1, 2)
                # Apply borders to the table
                self._apply_table_borders(ws, header_row, header_row + len(llm_df), 1, 2)
                
                for r in range(header_row + 1, header_row + 1 + len(llm_df)):
                    v = ws.cell(row=r, column=2)
                    v.alignment = Alignment(horizontal='right')
                    v.number_format = '#,##0.00'
                llm_last = header_row + len(llm_df)
                current_row = llm_last + 2

            # 3) Response Time table
            if rt:
                # Add title for Response Time table
                title_cell = ws.cell(row=current_row, column=1, value="Response Time (s)")
                title_cell.font = Font(bold=True, size=12)
                current_row += 1
                
                rt_df = pd.DataFrame([
                    ['Avg Time', rt.get('avg', 0.0)],
                    ['Min Time', rt.get('min', 0.0)],
                    ['Max Time', rt.get('max', 0.0)],
                    ['Median Time', rt.get('median', 0.0)],
                    ['Std Dev', rt.get('std', 0.0)],
                    ['Records Analyzed', rt.get('count', 0)],
                ], columns=['Metric', 'Value'])
                rt_df.to_excel(writer, sheet_name=ws.title, startrow=current_row-1, index=False)
                header_row = current_row
                # Apply enhanced header styling
                self._apply_header_styling(ws, header_row, 1, 2)
                # Apply borders to the table
                self._apply_table_borders(ws, header_row, header_row + len(rt_df), 1, 2)
                
                for r in range(header_row + 1, header_row + 1 + len(rt_df)):
                    v = ws.cell(row=r, column=2)
                    v.alignment = Alignment(horizontal='right')
                    v.number_format = '0.00'
                rt_last = header_row + len(rt_df)
                current_row = rt_last + 2

            # 4) Mode-wise and Process-wise tables when available
            m = data['metrics']
            # Mode-wise RT
            if m.get('rt_by_mode'):
                ws.cell(row=current_row, column=1, value='Response Time by Mode (s)').font = Font(bold=True)
                current_row += 1
                df = pd.DataFrame(m['rt_by_mode']) if isinstance(m['rt_by_mode'], list) else pd.DataFrame(m['rt_by_mode'])
                # Reorder columns if present
                cols = [c for c in ['effective_mode','mode_name','avg','p50','min','max','std','count'] if c in df.columns]
                df = df[cols]
                df.to_excel(writer, sheet_name=ws.title, startrow=current_row-1, index=False)
                # Apply numeric formats for RT columns (seconds)
                header_row = current_row
                last_row = header_row + len(df)
                headers = {cell.value: idx+1 for idx, cell in enumerate(ws[header_row])}
                for key in ['avg','p50','min','max','std']:
                    if key in headers:
                        col = headers[key]
                        for r in range(header_row + 1, last_row + 1):
                            cell = ws.cell(row=r, column=col)
                            if isinstance(cell.value, (int, float)):
                                cell.alignment = Alignment(horizontal='right')
                                cell.number_format = '0.00'
                for key in ['count','effective_mode']:
                    if key in headers:
                        col = headers[key]
                        for r in range(header_row + 1, last_row + 1):
                            cell = ws.cell(row=r, column=col)
                            if isinstance(cell.value, (int, float)):
                                cell.alignment = Alignment(horizontal='right')
                                cell.number_format = '0'
                current_row += len(df) + 2
            # Mode-wise Cost
            if m.get('cost_by_mode'):
                ws.cell(row=current_row, column=1, value='LLM Cost by Mode ($)').font = Font(bold=True)
                current_row += 1
                df = pd.DataFrame(m['cost_by_mode']) if isinstance(m['cost_by_mode'], list) else pd.DataFrame(m['cost_by_mode'])
                cols = [c for c in ['effective_mode','mode_name','avg','median','min','max','total','count'] if c in df.columns]
                df = df[cols]
                df.to_excel(writer, sheet_name=ws.title, startrow=current_row-1, index=False)
                # Apply numeric formats for currency columns
                header_row = current_row
                last_row = header_row + len(df)
                headers = {cell.value: idx+1 for idx, cell in enumerate(ws[header_row])}
                for key in ['avg','median','min','max','total']:
                    if key in headers:
                        col = headers[key]
                        for r in range(header_row + 1, last_row + 1):
                            cell = ws.cell(row=r, column=col)
                            if isinstance(cell.value, (int, float)):
                                cell.alignment = Alignment(horizontal='right')
                                cell.number_format = '#,##0.00'
                if 'count' in headers:
                    col = headers['count']
                    for r in range(header_row + 1, last_row + 1):
                        cell = ws.cell(row=r, column=col)
                        if isinstance(cell.value, (int, float)):
                            cell.alignment = Alignment(horizontal='right')
                            cell.number_format = '0'
                current_row += len(df) + 2
            # Mode-wise Failures
            if m.get('fail_by_mode'):
                ws.cell(row=current_row, column=1, value='Failure Rate by Mode').font = Font(bold=True)
                current_row += 1
                df = pd.DataFrame(m['fail_by_mode']) if isinstance(m['fail_by_mode'], list) else pd.DataFrame(m['fail_by_mode'])
                cols = [c for c in ['effective_mode','mode_name','error','info','total','failure_pct'] if c in df.columns]
                df = df[cols]
                df.to_excel(writer, sheet_name=ws.title, startrow=current_row-1, index=False)
                # Apply formats: counts as integers, failure_pct as percent
                header_row = current_row
                last_row = header_row + len(df)
                headers = {cell.value: idx+1 for idx, cell in enumerate(ws[header_row])}
                for key in ['error','info','total']:
                    if key in headers:
                        col = headers[key]
                        for r in range(header_row + 1, last_row + 1):
                            cell = ws.cell(row=r, column=col)
                            if isinstance(cell.value, (int, float)):
                                cell.alignment = Alignment(horizontal='right')
                                cell.number_format = '0'
                if 'failure_pct' in headers:
                    col = headers['failure_pct']
                    for r in range(header_row + 1, last_row + 1):
                        cell = ws.cell(row=r, column=col)
                        if isinstance(cell.value, (int, float)):
                            cell.alignment = Alignment(horizontal='right')
                            cell.number_format = '0.00%'
                current_row += len(df) + 2

            # Process-wise RT
            if m.get('rt_by_process'):
                ws.cell(row=current_row, column=1, value='Response Time by Process (s)').font = Font(bold=True)
                current_row += 1
                df = pd.DataFrame(m['rt_by_process'])
                cols = [c for c in ['process_name','avg','p50','min','max','std','count'] if c in df.columns]
                df = df[cols]
                df.to_excel(writer, sheet_name=ws.title, startrow=current_row-1, index=False)
                # Apply numeric formats (seconds)
                header_row = current_row
                last_row = header_row + len(df)
                headers = {cell.value: idx+1 for idx, cell in enumerate(ws[header_row])}
                for key in ['avg','p50','min','max','std']:
                    if key in headers:
                        col = headers[key]
                        for r in range(header_row + 1, last_row + 1):
                            cell = ws.cell(row=r, column=col)
                            if isinstance(cell.value, (int, float)):
                                cell.alignment = Alignment(horizontal='right')
                                cell.number_format = '0.00'
                if 'count' in headers:
                    col = headers['count']
                    for r in range(header_row + 1, last_row + 1):
                        cell = ws.cell(row=r, column=col)
                        if isinstance(cell.value, (int, float)):
                            cell.alignment = Alignment(horizontal='right')
                            cell.number_format = '0'
                current_row += len(df) + 2
            # Process-wise Cost
            if m.get('cost_by_process'):
                ws.cell(row=current_row, column=1, value='LLM Cost by Process ($)').font = Font(bold=True)
                current_row += 1
                df = pd.DataFrame(m['cost_by_process'])
                cols = [c for c in ['process_name','avg','median','min','max','total','count'] if c in df.columns]
                df = df[cols]
                df.to_excel(writer, sheet_name=ws.title, startrow=current_row-1, index=False)
                # Apply numeric formats (currency for costs)
                header_row = current_row
                last_row = header_row + len(df)
                headers = {cell.value: idx+1 for idx, cell in enumerate(ws[header_row])}
                for key in ['avg','median','min','max','total']:
                    if key in headers:
                        col = headers[key]
                        for r in range(header_row + 1, last_row + 1):
                            cell = ws.cell(row=r, column=col)
                            if isinstance(cell.value, (int, float)):
                                cell.alignment = Alignment(horizontal='right')
                                cell.number_format = '#,##0.00'
                for key in ['count','effective_mode']:
                    if key in headers:
                        col = headers[key]
                        for r in range(header_row + 1, last_row + 1):
                            cell = ws.cell(row=r, column=col)
                            if isinstance(cell.value, (int, float)):
                                cell.alignment = Alignment(horizontal='right')
                                cell.number_format = '0'
                current_row += len(df) + 2

            # Process-wise Failures
            if m.get('fail_by_process'):
                ws.cell(row=current_row, column=1, value='Failure Rate by Process').font = Font(bold=True)
                current_row += 1
                df = pd.DataFrame(m['fail_by_process'])
                cols = [c for c in ['process_name','error','info','total','failure_pct'] if c in df.columns]
                df = df[cols]
                df.to_excel(writer, sheet_name=ws.title, startrow=current_row-1, index=False)
                # Apply formats: counts as integers, failure_pct as percent
                header_row = current_row
                last_row = header_row + len(df)
                headers = {cell.value: idx+1 for idx, cell in enumerate(ws[header_row])}
                for key in ['error','info','total']:
                    if key in headers:
                        col = headers[key]
                        for r in range(header_row + 1, last_row + 1):
                            cell = ws.cell(row=r, column=col)
                            if isinstance(cell.value, (int, float)):
                                cell.alignment = Alignment(horizontal='right')
                                cell.number_format = '0'
                if 'failure_pct' in headers:
                    col = headers['failure_pct']
                    for r in range(header_row + 1, last_row + 1):
                        cell = ws.cell(row=r, column=col)
                        if isinstance(cell.value, (int, float)):
                            cell.alignment = Alignment(horizontal='right')
                            cell.number_format = '0.00%'
                current_row += len(df) + 2

            # Process × Mode RT
            if m.get('rt_by_process_mode'):
                ws.cell(row=current_row, column=1, value='Response Time by Process × Mode (s)').font = Font(bold=True)
                current_row += 1
                df = pd.DataFrame(m['rt_by_process_mode'])
                cols = [c for c in ['process_name','effective_mode','avg','p50','min','max','std','count'] if c in df.columns]
                df = df[cols]
                df.to_excel(writer, sheet_name=ws.title, startrow=current_row-1, index=False)
                current_row += len(df) + 2
            # Process × Mode Cost
            if m.get('cost_by_process_mode'):
                ws.cell(row=current_row, column=1, value='LLM Cost by Process × Mode ($)').font = Font(bold=True)
                current_row += 1
                df = pd.DataFrame(m['cost_by_process_mode'])
                cols = [c for c in ['process_name','effective_mode','avg','median','min','max','total','count'] if c in df.columns]
                df = df[cols]
                df.to_excel(writer, sheet_name=ws.title, startrow=current_row-1, index=False)
                current_row += len(df) + 2
            # Process × Mode Failures
            if m.get('fail_by_process_mode'):
                ws.cell(row=current_row, column=1, value='Failure Rate by Process × Mode').font = Font(bold=True)
                current_row += 1
                df = pd.DataFrame(m['fail_by_process_mode'])
                cols = [c for c in ['process_name','effective_mode','error','info','total','failure_pct'] if c in df.columns]
                df = df[cols]
                df.to_excel(writer, sheet_name=ws.title, startrow=current_row-1, index=False)
                current_row += len(df) + 2

            # 3) Charts block
            charts = data.get('charts', {})
            ordered = [
                'dauu_chart.png',
                'dau_chart.png',
                'mode_wise_dau_chart.png',
                'response_time_percentiles.png',
                'daily_response_time_range.png',
                'response_time_analysis.png',
                'error_categories_chart.png',
            ]
            for chart_file in ordered:
                if chart_file in charts:
                    try:
                        img = XLImage(charts[chart_file])
                        img.width = 720; img.height = 405
                        ws.add_image(img, f"A{current_row}")
                        current_row += 28
                    except Exception:
                        ws.cell(row=current_row, column=1, value=f"[Image not found: {charts[chart_file]}]")
                        current_row += 2

            # 4) Error Messages table (with derived Category column) - AFTER CHARTS
            msgs = data['metrics'].get('error_messages', {})
            if msgs:
                ws.cell(row=current_row, column=1, value='Error Messages').font = Font(bold=True, size=12)
                current_row += 1
                rows = []
                # Use pre-categorized mapping from individual analysis for consistency
                message_categories = data['metrics'].get('error_message_categories', {})
                for m, n in msgs.items():
                    # Use pre-categorized mapping if available, otherwise fall back to LLM service
                    cat = message_categories.get(m, self._categorize_error_message(m))
                    display_msg = m if len(m) <= 300 else m[:300]+"..."
                    rows.append([cat, display_msg, n])
                msg_df = pd.DataFrame(rows, columns=['Error Category', 'Error Message', 'Count'])
                # Sort by category then count desc
                msg_df.sort_values(by=['Error Category', 'Count'], ascending=[True, False], inplace=True)
                msg_df.to_excel(writer, sheet_name=ws.title, startrow=current_row-1, index=False)
                msg_header = current_row
                # Apply enhanced header styling
                self._apply_header_styling(ws, msg_header, 1, 3)
                # Apply borders to the table
                self._apply_table_borders(ws, msg_header, msg_header + len(msg_df), 1, 3)
                
                # Right-align counts (third column)
                for r in range(msg_header + 1, msg_header + 1 + len(msg_df)):
                    ws.cell(row=r, column=3).alignment = Alignment(horizontal='right')
                current_row = msg_header + len(msg_df) + 2

            # 5) Error Categories table - AFTER CHARTS
            cats = data['metrics'].get('error_categories', {})
            if cats:
                ws.cell(row=current_row, column=1, value='Error Categories').font = Font(bold=True, size=12)
                current_row += 1
                cat_df = pd.DataFrame([[c, n] for c, n in cats.items()], columns=['Error Category', 'Count'])
                cat_df.to_excel(writer, sheet_name=ws.title, startrow=current_row-1, index=False)
                cat_header = current_row
                # Apply enhanced header styling
                self._apply_header_styling(ws, cat_header, 1, 2)
                # Apply borders to the table
                self._apply_table_borders(ws, cat_header, cat_header + len(cat_df), 1, 2)
                
                for r in range(cat_header + 1, cat_header + 1 + len(cat_df)):
                    ws.cell(row=r, column=2).alignment = Alignment(horizontal='right')
                current_row = cat_header + len(cat_df) + 2

    # Removed By Service Overview as per request

    # Removed By Service Errors as per request

    def _create_index_sheet(self, writer):
        wb = writer.book
        # Create or get 'Link to other tabs'
        ws = wb.create_sheet('Link to other tabs', 0)
        
        # Title styling
        title_cell = ws.cell(row=1, column=1, value='Link to other tabs')
        title_cell.font = Font(bold=True, size=16, color='2F4F4F')
        title_cell.alignment = Alignment(horizontal='center')
        
        subtitle_cell = ws.cell(row=2, column=1, value='Click on any link below to jump to that sheet:')
        subtitle_cell.font = Font(size=12, italic=True, color='696969')
        
        sheets = [
            'Response Times', 'Success Rates', 'LLM Costs',
            'Error Categories', 'Detailed Error Messages', 'Charts'
        ]
        # Include per-service sheets if any
        if hasattr(self, '_service_sheet_names'):
            sheets.extend(self._service_sheet_names)
        
        row = 4
        for name in sheets:
            if name in wb.sheetnames:
                cell = ws.cell(row=row, column=1)
                cell.value = f"=HYPERLINK(\"#'{name}'!A1\",\"{name}\")"
                cell.font = Font(size=11, color='0066CC', underline='single')
                cell.alignment = Alignment(horizontal='left')
                row += 1
        
        # Auto-adjust column width
        ws.column_dimensions['A'].width = 30

    def _apply_table_borders(self, ws, start_row, end_row, start_col, end_col):
        """Apply borders to a table range"""
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in range(start_row, end_row + 1):
            for col in range(start_col, end_col + 1):
                cell = ws.cell(row=row, column=col)
                cell.border = thin_border

    def _apply_header_styling(self, ws, row, start_col, end_col):
        """Apply header styling to a row"""
        header_fill = PatternFill(start_color='E6E6FA', end_color='E6E6FA', fill_type='solid')
        
        for col in range(start_col, end_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = Font(bold=True, size=11)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')

    # --- ALL PDF GENERATION CODE REMAINS THE SAME AS THE PREVIOUS POLISHED VERSION ---
    def generate_pdf_report(self, all_data: Dict) -> bool:
        try:
            today = datetime.now().strftime('%Y%m%d_%H%M')
            pdf_path = f"{self.reports_dir}/analysis_report_{today}.pdf"
            with PdfPages(pdf_path) as pdf:
                self._create_pdf_title(pdf)
                for file_name, data in all_data.items():
                    self._create_pdf_combined_metrics_table(pdf, file_name, data)
                    if data['metrics'].get('error_categories') or data['metrics'].get('error_messages'):
                        self._create_pdf_error_tables(pdf, file_name, data)
                    # Process-wise tables page when present
                    self._create_pdf_process_tables(pdf, file_name, data)
                    # Mode-wise tables page when present
                    self._create_pdf_mode_tables(pdf, file_name, data)
                    self._create_pdf_document_charts(pdf, file_name, data)
            print(f"✅ PDF report: {pdf_path}")
            return True
        except Exception as e:
            print(f"❌ PDF generation failed: {e}")
            traceback.print_exc()
            return False
    
    def _save_page_to_pdf(self, pdf, fig):
        # Add minimal footer: page number at bottom-right to avoid overlap
        footer_text = f"{self._pdf_page_num + 1}"
        fig.text(0.99, 0.015, footer_text, ha='right', va='center', fontsize=9, color='gray')
        pdf.savefig(fig, bbox_inches=None, pad_inches=0.5)
        self._pdf_page_num += 1
        plt.close(fig)

    def _create_pdf_title(self, pdf):
        fig = plt.figure(figsize=self.A4_SIZE_INCHES)
        fig.text(0.5, 0.6, 'DataDog Analysis Report', ha='center', va='center', fontsize=28, weight='bold')
        fig.text(0.5, 0.45, datetime.now().strftime('%B %d, %Y'), ha='center', va='center', fontsize=16)
        plt.axis('off')
        self._save_page_to_pdf(pdf, fig)
    
    def _create_pdf_combined_metrics_table(self, pdf, file_name: str, data: Dict):
        fig = plt.figure(figsize=self.A4_SIZE_INCHES)
        fig.text(0.5, 0.95, f'Metrics Summary: {file_name}', ha='center', va='center', fontsize=18, weight='bold')
        current_y = 0.90
        rt = data['metrics'].get('response_time', {})
        rt_data = [
            ['Average Time', f"{rt.get('avg', 0):.2f}"], ['Min Time', f"{rt.get('min', 0):.2f}"],
            ['Max Time', f"{rt.get('max', 0):.2f}"], ['Median Time', f"{rt.get('median', 0):.2f}"]
        ]
        axis_height = 0.05 + len(rt_data) * 0.035
        axis_bottom = current_y - axis_height
        ax1 = fig.add_axes([0.1, axis_bottom, 0.8, axis_height])
        ax1.set_title('Response Time Metrics', fontsize=12, weight='bold', pad=10)
        ax1.axis('off')
        self._render_table(ax1, rt_data, ['Metric', 'Value'])
        current_y = axis_bottom - 0.04
        st = data['metrics'].get('status', {})
        status_data = [
            ['Success', f"{st.get('success_count', 0):,}", f"{st.get('success_rate', 0):.2f}"],
            ['Error', f"{st.get('error_count', 0):,}", f"{st.get('error_rate', 0):.2f}"],
            ['Total', f"{st.get('total', 0):,}", '100.00']
        ]
        axis_height = 0.05 + len(status_data) * 0.035
        axis_bottom = current_y - axis_height
        ax2 = fig.add_axes([0.1, axis_bottom, 0.8, axis_height])
        ax2.set_title('Success & Failure Metrics', fontsize=12, weight='bold', pad=10)
        ax2.axis('off')
        self._render_table(ax2, status_data, ['Status', 'Count', 'Rate'], col_widths=[0.4, 0.3, 0.3])
        current_y = axis_bottom - 0.04
        if 'llm_cost' in data['metrics']:
            cost = data['metrics'].get('llm_cost', {})
            cost_data = [
                ['Average Cost', f"{cost.get('avg', 0):.4f}"], ['Min Cost', f"{cost.get('min', 0):.4f}"],
                ['Max Cost', f"{cost.get('max', 0):.4f}"], ['Median Cost', f"{cost.get('median', 0):.4f}"],
                ['Total Cost', f"{cost.get('total', 0):.2f}"]
            ]
            axis_height = 0.05 + len(cost_data) * 0.035
            axis_bottom = current_y - axis_height
            ax3 = fig.add_axes([0.1, axis_bottom, 0.8, axis_height])
            ax3.set_title('LLM Cost Metrics', fontsize=12, weight='bold', pad=10)
            ax3.axis('off')
            self._render_table(ax3, cost_data, ['Metric', 'Value'])
        self._save_page_to_pdf(pdf, fig)

    def _create_pdf_error_tables(self, pdf, file_name: str, data: Dict):
        fig = plt.figure(figsize=self.A4_SIZE_INCHES)
        fig.text(0.5, 0.95, f'Error Analysis: {file_name}', ha='center', va='center', fontsize=18, weight='bold')
        current_y = 0.90
        has_categories = 'error_categories' in data['metrics'] and data['metrics']['error_categories']
        has_messages = 'error_messages' in data['metrics'] and data['metrics']['error_messages']
        # Build Category → Messages table if messages exist
        if has_messages:
            rows = []
            # Use pre-categorized mapping from individual analysis for consistency
            message_categories = data['metrics'].get('error_message_categories', {})
            for msg, count in data['metrics']['error_messages'].items():
                # Use pre-categorized mapping if available, otherwise fall back to LLM service
                cat = message_categories.get(msg, self._categorize_error_message(msg))
                rows.append([cat, msg, f"{count:,}"])
            if rows:
                desired_height = 0.05 + len(rows) * 0.03
                avail = current_y - 0.12
                if avail <= 0.12:
                    # New page
                    self._save_page_to_pdf(pdf, fig)
                    fig = plt.figure(figsize=self.A4_SIZE_INCHES)
                    fig.text(0.5, 0.95, f'Error Analysis: {file_name}', ha='center', va='center', fontsize=18, weight='bold')
                    current_y = 0.90
                    avail = current_y - 0.12
                axis_height = min(desired_height, max(0.12, avail))
                axis_bottom = 0.08
                ax0 = fig.add_axes([0.1, axis_bottom, 0.8, axis_height])
                ax0.set_title('Error Category → Messages', fontsize=12, weight='bold', pad=10)
                ax0.axis('off')
                rows_sorted = sorted(rows, key=lambda x: (x[0], -int(x[2].replace(',', ''))))
                # Give message column more width to avoid overlap
                self._render_table(ax0, rows_sorted, ['Category', 'Message', 'Count'], col_widths=[0.22, 0.63, 0.15])
                current_y = axis_bottom - 0.04
        if has_categories:
            cats = data['metrics']['error_categories']
            cat_data = [[cat, f"{count:,}"] for cat, count in cats.items()]
            desired_height = 0.05 + len(cat_data) * 0.035
            avail = current_y - 0.12
            if avail <= 0.12:
                self._save_page_to_pdf(pdf, fig)
                fig = plt.figure(figsize=self.A4_SIZE_INCHES)
                fig.text(0.5, 0.95, f'Error Analysis: {file_name}', ha='center', va='center', fontsize=18, weight='bold')
                current_y = 0.90
                avail = current_y - 0.12
            axis_height = min(desired_height, max(0.12, avail))
            axis_bottom = 0.08
            ax1 = fig.add_axes([0.1, axis_bottom, 0.8, axis_height])
            ax1.set_title('Error Categories', fontsize=12, weight='bold', pad=10)
            ax1.axis('off')
            self._render_table(ax1, cat_data, ['Error Category', 'Count'])
            current_y = axis_bottom - 0.04
        if has_messages:
            # Messages as wrapped text (may spill to new page if not enough space)
            msgs = data['metrics']['error_messages']
            avail = current_y - 0.12
            if avail <= 0.12:
                self._save_page_to_pdf(pdf, fig)
                fig = plt.figure(figsize=self.A4_SIZE_INCHES)
                fig.text(0.5, 0.95, f'Error Analysis: {file_name}', ha='center', va='center', fontsize=18, weight='bold')
                current_y = 0.90
                avail = current_y - 0.12
            axis_height = max(0.12, avail)
            axis_bottom = 0.08
            ax2 = fig.add_axes([0.1, axis_bottom, 0.8, axis_height])
            ax2.set_title('Detailed Error Messages', fontsize=12, weight='bold', pad=10)
            ax2.axis('off')
            y = 0.92
            line_height = 0.05
            for msg, count in msgs.items():
                wrapped = textwrap.wrap(f"• {msg} (Count: {count:,})", width=90)
                for line in wrapped:
                    if y < 0.05:
                        break
                    ax2.text(0.0, y, line, fontsize=10, ha='left', va='top')
                    y -= line_height
        self._save_page_to_pdf(pdf, fig)

    def _create_pdf_process_tables(self, pdf, file_name: str, data: Dict):
        """Create a page with process-wise RT, LLM cost, and failure tables if available."""
        m = data.get('metrics', {})
        has_rt = bool(m.get('rt_by_process'))
        has_cost = bool(m.get('cost_by_process'))
        has_fail = bool(m.get('fail_by_process'))
        if not (has_rt or has_cost or has_fail):
            return
        fig = plt.figure(figsize=self.A4_SIZE_INCHES)
        fig.text(0.5, 0.95, f'Process-wise Metrics: {file_name}', ha='center', va='center', fontsize=18, weight='bold')
        current_y = 0.90
        blocks = []
        if has_rt:
            rows = []
            for r in m['rt_by_process']:
                rows.append([
                    r.get('process_name',''), f"{r.get('avg',0):.2f}", f"{r.get('p50',0):.2f}", f"{r.get('min',0):.2f}", f"{r.get('max',0):.2f}", f"{r.get('std',0):.2f}", f"{r.get('count',0):,}"
                ])
            blocks.append(('Response Time by Process', ['Process Name','Avg','P50','Min','Max','Std','N'], rows))
        if has_cost:
            rows = []
            for r in m['cost_by_process']:
                rows.append([
                    r.get('process_name',''), f"{r.get('avg',0):.4f}", f"{r.get('median',0):.4f}", f"{r.get('min',0):.4f}", f"{r.get('max',0):.4f}", f"{r.get('total',0):.2f}", f"{r.get('count',0):,}"
                ])
            blocks.append(('LLM Cost by Process', ['Process Name','Avg','Median','Min','Max','Total','N'], rows))
        if has_fail:
            rows = []
            for r in m['fail_by_process']:
                rows.append([
                    r.get('process_name',''), f"{r.get('error',0):,}", f"{r.get('info',0):,}", f"{r.get('total',0):,}", f"{r.get('failure_pct',0):.2f}"
                ])
            blocks.append(('Failure Rate by Process', ['Process Name','Error','Success (Info)','Total','Failure Rate'], rows))

        for title, headers, rows in blocks:
            axis_height = 0.05 + max(1, len(rows)) * 0.035
            axis_bottom = current_y - axis_height
            ax = fig.add_axes([0.05, axis_bottom, 0.90, axis_height])
            ax.set_title(title, fontsize=12, weight='bold', pad=10)
            ax.axis('off')
            # Adjust widths for long process names
            widths = [0.35] + [ (0.65 / (len(headers)-1)) for _ in headers[1:] ]
            self._render_table(ax, rows, headers, col_widths=widths)
            current_y = axis_bottom - 0.04
            if current_y < 0.15:
                break
        self._save_page_to_pdf(pdf, fig)

    def _create_pdf_mode_tables(self, pdf, file_name: str, data: Dict):
        """Create a page with mode-wise RT, LLM cost, and failure tables if available."""
        m = data.get('metrics', {})
        has_rt = bool(m.get('rt_by_mode'))
        has_cost = bool(m.get('cost_by_mode'))
        has_fail = bool(m.get('fail_by_mode'))
        if not (has_rt or has_cost or has_fail):
            return
        fig = plt.figure(figsize=self.A4_SIZE_INCHES)
        fig.text(0.5, 0.95, f'Mode-wise Metrics: {file_name}', ha='center', va='center', fontsize=18, weight='bold')
        current_y = 0.90
        # Layout up to three stacked tables
        blocks = []
        if has_rt:
            rt_rows = []
            for r in m['rt_by_mode']:
                rt_rows.append([
                    r.get('effective_mode',''), r.get('mode_name',''),
                    f"{r.get('avg',0):.2f}", f"{r.get('p50',0):.2f}", f"{r.get('min',0):.2f}", f"{r.get('max',0):.2f}",
                    f"{r.get('std',0):.2f}", f"{r.get('count',0):,}"
                ])
            blocks.append(('Response Time by Mode', ['Mode','Name','Avg','P50','Min','Max','Std','N'], rt_rows))
        if has_cost:
            cost_rows = []
            for r in m['cost_by_mode']:
                cost_rows.append([
                    r.get('effective_mode',''), r.get('mode_name',''),
                    f"{r.get('avg',0):.4f}", f"{r.get('median',0):.4f}", f"{r.get('min',0):.4f}", f"{r.get('max',0):.4f}",
                    f"{r.get('total',0):.2f}", f"{r.get('count',0):,}"
                ])
            blocks.append(('LLM Cost by Mode', ['Mode','Name','Avg','Median','Min','Max','Total','N'], cost_rows))
        if has_fail:
            fail_rows = []
            for r in m['fail_by_mode']:
                fail_rows.append([
                    r.get('effective_mode',''), r.get('mode_name',''),
                    f"{r.get('error',0):,}", f"{r.get('info',0):,}", f"{r.get('total',0):,}", f"{r.get('failure_pct',0):.2f}"
                ])
            blocks.append(('Failure Rate by Mode', ['Mode','Name','Error','Success (Info)','Total','Failure Rate'], fail_rows))

        # Render blocks
        for title, headers, rows in blocks:
            # Heuristic height per block
            axis_height = 0.05 + max(1, len(rows)) * 0.035
            axis_bottom = current_y - axis_height
            ax = fig.add_axes([0.1, axis_bottom, 0.8, axis_height])
            ax.set_title(title, fontsize=12, weight='bold', pad=10)
            ax.axis('off')
            self._render_table(ax, rows, headers, col_widths=[0.12,0.28,0.12,0.12,0.12,0.12,0.12,0.10][:len(headers)])
            current_y = axis_bottom - 0.04
            if current_y < 0.15:
                break  # avoid overflow; future improvement: paginate if needed
        self._save_page_to_pdf(pdf, fig)
    
    def _render_table(self, ax, data, headers, col_widths=None):
        if col_widths is None:
            # Default: favor text-heavy second column if present
            col_widths = [0.22] + [0.63] + [0.15] if len(headers) == 3 else [0.6, 0.4]
        table = ax.table(cellText=data, colLabels=headers, loc='upper center', cellLoc='left', colWidths=col_widths)
        table.auto_set_font_size(False); table.set_fontsize(10); table.scale(1, 1.8)
        for (i, j), cell in table.get_celld().items():
            cell.set_edgecolor('lightgray')
            if i == 0:
                cell.set_text_props(weight='bold', color='white', ha='center')
                cell.set_facecolor('#2c5a8c')
            else:
                # Zebra striping for data rows
                if i % 2 == 0:
                    cell.set_facecolor('#f7f9fc')
                else:
                    cell.set_facecolor('white')
                cell.set_text_props(va='top')
                # Left-align text columns except last numeric Count column
                if headers and headers[-1].lower() == 'count' and j == len(headers)-1:
                    cell.set_text_props(ha='right')
                else:
                    cell.set_text_props(ha='left')
    
    def _create_pdf_document_charts(self, pdf, file_name: str, data: Dict):
        charts = data.get('charts', {})
        # Combine DAU and DAUU on one page when available
        dau_path = charts.get('dau_chart.png')
        dauu_path = charts.get('dauu_chart.png')
        if dau_path or dauu_path:
            self._create_stacked_chart_page(pdf, file_name,
                                            top_image_path=dau_path,
                                            bottom_image_path=dauu_path,
                                            top_title='Daily Active Users (DAU)',
                                            bottom_title='Daily Active Unique Users (DAUU)')
        # Mode-wise DAU page when available
        if 'mode_wise_dau_chart.png' in charts:
            self._create_chart_page(pdf, file_name, charts['mode_wise_dau_chart.png'], 'Mode-wise DAU')
        # Render remaining charts (without descriptions), one per page
        # Combine response-time charts row-wise on a single page
        rt_top = charts.get('response_time_percentiles.png')
        rt_bottom = charts.get('daily_response_time_range.png')
        if rt_top or rt_bottom:
            self._create_stacked_chart_page(pdf, file_name,
                                            top_image_path=rt_top,
                                            bottom_image_path=rt_bottom,
                                            top_title='Response Time Percentiles',
                                            bottom_title='Daily Response Time Range & Avg')
        # Keep the analysis dashboard as a single concise page if present
        if 'response_time_analysis.png' in charts:
            self._create_chart_page(pdf, file_name, charts['response_time_analysis.png'], 'Response Time Analysis')

    def _create_chart_page(self, pdf, file_name, image_path, title):
        fig = plt.figure(figsize=self.A4_SIZE_INCHES)
        fig.text(0.5, 0.95, f'{file_name}\n{title}', ha='center', va='top', fontsize=16, weight='bold', wrap=True)
        ax_img = fig.add_axes([0.05, 0.08, 0.9, 0.80])
        try:
            with Image.open(image_path) as img:
                ax_img.imshow(img)
        except FileNotFoundError:
            ax_img.text(0.5, 0.5, 'Chart image not found.', ha='center', va='center', color='red')
        ax_img.axis('off')
        self._save_page_to_pdf(pdf, fig)

    def _create_dual_chart_page(self, pdf, file_name, left_image_path, right_image_path, left_title='Left', right_title='Right'):
        fig = plt.figure(figsize=self.A4_SIZE_INCHES)
        fig.text(0.5, 0.95, f'{file_name}\n{left_title} | {right_title}', ha='center', va='top', fontsize=16, weight='bold', wrap=True)
        # Left image area
        ax_left = fig.add_axes([0.05, 0.08, 0.425, 0.80])
        ax_left.axis('off')
        if left_image_path:
            try:
                with Image.open(left_image_path) as img:
                    ax_left.imshow(img)
            except FileNotFoundError:
                ax_left.text(0.5, 0.5, 'Chart image not found.', ha='center', va='center', color='red')
        # Right image area
        ax_right = fig.add_axes([0.525, 0.08, 0.425, 0.80])
        ax_right.axis('off')
        if right_image_path:
            try:
                with Image.open(right_image_path) as img:
                    ax_right.imshow(img)
            except FileNotFoundError:
                ax_right.text(0.5, 0.5, 'Chart image not found.', ha='center', va='center', color='red')
        self._save_page_to_pdf(pdf, fig)

    def _create_stacked_chart_page(self, pdf, file_name, top_image_path, bottom_image_path, top_title='Top', bottom_title='Bottom'):
        fig = plt.figure(figsize=self.A4_SIZE_INCHES)
        fig.text(0.5, 0.95, f'{file_name}\n{top_title} / {bottom_title}', ha='center', va='top', fontsize=16, weight='bold', wrap=True)
        # Top image area
        ax_top = fig.add_axes([0.05, 0.52, 0.90, 0.35])
        ax_top.axis('off')
        if top_image_path:
            try:
                with Image.open(top_image_path) as img:
                    ax_top.imshow(img)
            except FileNotFoundError:
                ax_top.text(0.5, 0.5, 'Chart image not found.', ha='center', va='center', color='red')
        # Bottom image area
        ax_bottom = fig.add_axes([0.05, 0.08, 0.90, 0.35])
        ax_bottom.axis('off')
        if bottom_image_path:
            try:
                with Image.open(bottom_image_path) as img:
                    ax_bottom.imshow(img)
            except FileNotFoundError:
                ax_bottom.text(0.5, 0.5, 'Chart image not found.', ha='center', va='center', color='red')
        self._save_page_to_pdf(pdf, fig)
    
    def generate_reports(self) -> bool:
        """Generate Excel report only"""
        print("🚀 Generating final, polished combined reports (regex-only parsing)...")
        all_data = self.collect_data()
        if not all_data:
            print("❌ No data found!")
            return False
        excel_success = self.generate_excel_report(all_data)
        if excel_success:
            print("\n🎉 Report generated successfully!")
        else:
            print("\n⚠️ Report failed to generate.")
        return excel_success

def main():
    generator = FinalPolishedCombinedReport()
    success = generator.generate_reports()
    return 0 if success else 1

if __name__ == "__main__":
    sys.exit(main())